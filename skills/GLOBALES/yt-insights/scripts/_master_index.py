"""Append a row to MASTER_INDEX.xlsx."""
import re
from pathlib import Path
from openpyxl import load_workbook


def _extract_top_3_titles(briefs_md: str) -> str:
    """Pull the 3 idea titles from the briefs."""
    titles = []
    for m in re.finditer(r"^###\s+🏆\s+IDEA\s*#?\d+\s*[—-]\s*(.+)$", briefs_md, re.MULTILINE):
        titles.append(m.group(1).strip())
    return " | ".join(titles[:3])


def _extract_tema(insights_md: str) -> str:
    """Pull the 'Tema central' line from insights."""
    m = re.search(r"\*\*Tema central:\*\*\s*(.+)", insights_md)
    if m:
        return m.group(1).strip().rstrip(".")
    return ""


def append_to_master_index(xlsx_path: Path, folder: Path, meta: dict, insights_md: str, briefs_md: str):
    """Upsert: if slug already exists, update the row in place. Else append."""
    wb = load_workbook(xlsx_path)
    ws = wb["videos"]

    slug = folder.name
    duration_min = (meta.get("duration_seconds") or 0) // 60
    tema = _extract_tema(insights_md)
    top3 = _extract_top_3_titles(briefs_md)

    row = [
        slug,
        meta.get("title", ""),
        meta.get("channel", ""),
        meta.get("url", ""),
        duration_min,
        meta.get("upload_date", ""),
        meta.get("extracted_at", ""),
        tema,
        "processed",
        top3,
        str(folder),
        str(folder / "insights.pdf"),
        str(folder / "content_briefs.pdf"),
        str(folder / "_RECAP.pdf"),
    ]

    # Find existing row by slug (column 1)
    target_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == slug:
            target_row = r
            break

    if target_row is None:
        target_row = ws.max_row + 1

    for col, val in enumerate(row, 1):
        ws.cell(row=target_row, column=col, value=val)

    wb.save(xlsx_path)
