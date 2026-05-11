"""One-shot: add Research Log + Asset Index sheets to the master xlsx if missing."""
import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# Helper compartido del repo
_HERE = Path(__file__).resolve()
for _i in range(2, 7):
    _candidate = _HERE.parents[_i] / "tools"
    if (_candidate / "spekgen_paths.py").exists():
        sys.path.insert(0, str(_candidate))
        break

try:
    from spekgen_paths import client_dir
    XLSX = client_dir("F24") / "F24 - 02. PRODUCTOS" / "00_50_PRIORITARIOS_MARVELSA.xlsx"
except ImportError:
    if os.environ.get("F24_PRODUCTS_DIR"):
        XLSX = Path(os.environ["F24_PRODUCTS_DIR"]) / "00_50_PRIORITARIOS_MARVELSA.xlsx"
    elif os.environ.get("SPEKGEN_ROOT"):
        XLSX = Path(os.environ["SPEKGEN_ROOT"]) / "F24- FERRE24" / "F24 - 02. PRODUCTOS" / "00_50_PRIORITARIOS_MARVELSA.xlsx"
    else:
        raise RuntimeError("No encontré tools/spekgen_paths.py ni SPEKGEN_ROOT/F24_PRODUCTS_DIR.")

RESEARCH_LOG_COLS = [
    "SKU", "Campo", "Valor", "Fuente URL", "Confidence", "Agente", "Fecha", "Notas"
]
ASSET_INDEX_COLS = [
    "SKU", "Tipo", "Filename", "Ruta relativa", "Size KB", "Source URL", "Fecha", "Notas"
]

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")

def ensure_sheet(wb, name, cols):
    if name in wb.sheetnames:
        print(f"  [=] '{name}' already exists, skipping")
        return wb[name]
    ws = wb.create_sheet(name)
    for i, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(14, len(col) + 2)
    ws.freeze_panes = "A2"
    print(f"  [+] created '{name}' with {len(cols)} cols")
    return ws

def main():
    wb = openpyxl.load_workbook(XLSX)
    print(f"Loaded {XLSX.name}, existing sheets: {wb.sheetnames}")
    ensure_sheet(wb, "Research Log", RESEARCH_LOG_COLS)
    ensure_sheet(wb, "Asset Index", ASSET_INDEX_COLS)
    wb.save(XLSX)
    print(f"Saved. Final sheets: {wb.sheetnames}")

if __name__ == "__main__":
    main()
