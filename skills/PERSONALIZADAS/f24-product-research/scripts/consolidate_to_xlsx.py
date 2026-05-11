"""Consolida los 4 JSON outputs de los agentes research → Research Log + Asset Index + actualiza Checklist Recursos.

Uso:
    python3 _consolidate_to_xlsx.py <sku> <product_folder_name>

Ejemplo:
    python3 _consolidate_to_xlsx.py HP5.5N "Hidrolavadora a gasolina 5.5hp"
"""
import json
import os
import sys
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font

# === Resolver ROOT via helper compartido del repo (tools/spekgen_paths.py) ===
# Script live en skills/PERSONALIZADAS/f24-product-research/scripts/, helper en tools/.
_HERE = Path(__file__).resolve()
for _i in range(2, 7):
    _candidate_tools = _HERE.parents[_i] / "tools"
    if (_candidate_tools / "spekgen_paths.py").exists():
        sys.path.insert(0, str(_candidate_tools))
        break

try:
    from spekgen_paths import client_dir
    ROOT = client_dir("F24") / "F24 - 02. PRODUCTOS"
except ImportError:
    if os.environ.get("F24_PRODUCTS_DIR"):
        ROOT = Path(os.environ["F24_PRODUCTS_DIR"])
    elif os.environ.get("SPEKGEN_ROOT"):
        ROOT = Path(os.environ["SPEKGEN_ROOT"]) / "F24- FERRE24" / "F24 - 02. PRODUCTOS"
    else:
        raise RuntimeError(
            "No encontré tools/spekgen_paths.py ni SPEKGEN_ROOT/F24_PRODUCTS_DIR env var. "
            "Asegúrate de que el repo spekgen-os esté clonado completo (con tools/spekgen_paths.py)."
        )

XLSX = ROOT / "00_50_PRIORITARIOS_MARVELSA.xlsx"
TODAY = datetime.now().strftime("%Y-%m-%d")

FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_PARTIAL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_MISS = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


# === Mapping: campo del checklist → cómo derivar valor/status desde los findings ===

# Columns en Checklist Recursos (de la fila header — cols 1-29):
# 1.Código  2.Producto  3.URL Marvelsa  4.%Completo  5.Notas
# A. MEDIA: 6.Hero  7.Secundarias3+  8.Lifestyle  9.Video
# B. DOCS: 10.Ficha  11.Manual  12.Exploded
# C. SPECS: 13.Marca+modelo  14.Motor  15.Operativas  16.Físicas  17.Garantía  18.NOM  19.Origen
# D. COMERCIAL: 20.CategShopify  21.Tags  22.PrecioDuarte  23.Inventario  24.DescCorta  25.DescLarga  26.Bullets  27.SEO
# E. LOGÍSTICA: 28.Peso/dim embarque  29.Frágil
CHECKLIST_FIELDS = [
    ("hero", "A. MEDIA / Hero", 6),
    ("secundarias", "A. MEDIA / Secundarias 3+", 7),
    ("lifestyle", "A. MEDIA / Lifestyle", 8),
    ("video", "A. MEDIA / Video demo", 9),
    ("ficha_tecnica", "B. DOCS / Ficha técnica PDF", 10),
    ("manual", "B. DOCS / Manual de uso PDF", 11),
    ("exploded", "B. DOCS / Diagrama exploded", 12),
    ("marca_modelo", "C. SPECS / Marca + modelo", 13),
    ("motor", "C. SPECS / Motor", 14),
    ("operativas", "C. SPECS / Operativas", 15),
    ("fisicas", "C. SPECS / Físicas", 16),
    ("garantia", "C. SPECS / Garantía", 17),
    ("nom", "C. SPECS / NOM", 18),
    ("pais_origen", "C. SPECS / País origen", 19),
    ("categoria_shopify", "D. COMERCIAL / Categoría Shopify", 20),
    ("tags", "D. COMERCIAL / Tags/keywords", 21),
    ("precio_duarte", "D. COMERCIAL / Precio venta Duarte", 22),
    ("inventario", "D. COMERCIAL / Inventario/bodega", 23),
    ("desc_corta", "D. COMERCIAL / Desc corta", 24),
    ("desc_larga", "D. COMERCIAL / Desc larga", 25),
    ("bullets", "D. COMERCIAL / Bullets venta", 26),
    ("seo", "D. COMERCIAL / SEO meta+slug", 27),
    ("peso_dim_embarque", "E. LOGÍSTICA / Peso/dim embarque", 28),
    ("fragil", "E. LOGÍSTICA / Frágil/flete especial", 29),
]


def load_findings(folder: Path):
    """Carga los 5 archivos JSON. specs.json y comercial.json son outputs principales."""
    out = {}
    paths = {
        "media": folder / "01. MEDIA" / "_findings.json",
        "docs": folder / "02. DOCS" / "_findings.json",
        "specs": folder / "03. SPECS" / "specs.json",
        "comercial": folder / "04. COMERCIAL" / "comercial.json",
    }
    for k, p in paths.items():
        if p.exists():
            out[k] = json.loads(p.read_text(encoding="utf-8"))
        else:
            print(f"  WARN: {p} not found")
            out[k] = None
    return out


def derive_field_state(field_key: str, f: dict, sku_row: dict) -> tuple[str, str, str, str, str]:
    """Returns (checklist_flag, valor, fuente_url, confidence, notas) per campo."""
    media = f.get("media") or {}
    docs = f.get("docs") or {}
    specs = f.get("specs") or {}
    com = f.get("comercial") or {}

    def flag(present, partial=False):
        if present and not partial:
            return "✓ Listo"
        if partial:
            return "⚠ Parcial"
        return "✗ Falta"

    # MEDIA
    if field_key == "hero":
        fld = (media.get("fields") or {}).get("hero", {})
        files = fld.get("files", [])
        if files:
            f0 = files[0]
            return (flag(True), f0.get("filename"), f0.get("source_url"), f0.get("confidence"), "")
        return (flag(False), "", "", "", "no encontrado")

    if field_key == "secundarias":
        fld = (media.get("fields") or {}).get("secundarias", {})
        files = fld.get("files", [])
        n = len(files)
        st = fld.get("status", "not_found")
        if n >= 3:
            return ("✓ Listo", f"{n} imágenes", files[0].get("source_url", ""), "alta", f"{n} secundarias")
        if n > 0:
            return ("⚠ Parcial", f"{n} imágenes", files[0].get("source_url", ""), "media", f"solo {n} de 3 mínimas")
        return ("✗ Falta", "", "", "", "")

    if field_key == "lifestyle":
        fld = (media.get("fields") or {}).get("lifestyle", {})
        files = fld.get("files", [])
        if files:
            f0 = files[0]
            conf = f0.get("confidence", "media")
            return (flag(True, partial=(conf == "baja")), f0.get("filename"), f0.get("source_url"), conf, f0.get("notas", ""))
        return ("✗ Falta", "", "", "", "")

    if field_key == "video":
        fld = (media.get("fields") or {}).get("video", {})
        urls = fld.get("urls", [])
        if urls:
            return ("✓ Listo" if len(urls) >= 1 else "⚠ Parcial", f"{len(urls)} URLs", urls[0].get("url"), "alta", urls[0].get("title", ""))
        return ("✗ Falta", "", "", "", "")

    # DOCS
    if field_key in ("ficha_tecnica", "manual", "exploded"):
        fld = (docs.get("fields") or {}).get(field_key, {})
        st = fld.get("status", "not_found")
        if st == "found":
            return ("✓ Listo", fld.get("filename", ""), fld.get("source_url", ""), fld.get("confidence", "alta"), fld.get("notas", ""))
        if st == "found_in_manual":
            return ("⚠ Parcial", "dentro del manual", fld.get("source_url", ""), "media", "ver páginas referenciadas")
        return ("✗ Falta", "", "", "", fld.get("notas", "not_found"))

    # SPECS
    if field_key == "marca_modelo":
        marca = specs.get("marca")
        modelo = specs.get("modelo")
        if marca and modelo:
            return ("✓ Listo", f"{marca} {modelo}", (specs.get("motor") or {}).get("fuentes", [""])[0], "alta", "")
        if marca or modelo:
            return ("⚠ Parcial", f"{marca or '?'} {modelo or '?'}", "", "media", "")
        return ("✗ Falta", "", "", "", "")

    if field_key == "motor":
        m = specs.get("motor") or {}
        # consider complete if HP + tipo + combustible + rpm
        if m.get("potencia_hp") and m.get("tipo") and m.get("combustible"):
            partial = m.get("cilindrada_cc") is None or m.get("tanque_litros") is None
            return ("⚠ Parcial" if partial else "✓ Listo",
                    f"{m['potencia_hp']}HP {m['tipo']} {m['combustible']}",
                    (m.get("fuentes") or [""])[0],
                    m.get("confidence", "media"),
                    "faltan cc y tanque" if partial else "")
        return ("✗ Falta", "", "", "", "")

    if field_key == "operativas":
        o = specs.get("operativas") or {}
        if o.get("presion_psi") and o.get("caudal_lpm"):
            partial = o.get("manguera_metros") is None
            return ("⚠ Parcial" if partial else "✓ Listo",
                    f"{o['presion_psi']} PSI / {o['caudal_lpm']} L/min",
                    (o.get("fuentes") or [""])[0],
                    o.get("confidence", "media"),
                    "falta longitud manguera" if partial else "")
        return ("✗ Falta", "", "", "", "")

    if field_key == "fisicas":
        fi = specs.get("fisicas") or {}
        if fi.get("peso_kg") and fi.get("dim_largo_cm"):
            partial = fi.get("color") is None
            return ("⚠ Parcial" if partial else "✓ Listo",
                    f"{fi['peso_kg']}kg / {fi.get('dim_largo_cm')}x{fi.get('dim_ancho_cm')}x{fi.get('dim_alto_cm')}cm",
                    (fi.get("fuentes") or [""])[0],
                    fi.get("confidence", "media"),
                    fi.get("dim_nota", ""))
        return ("✗ Falta", "", "", "", "")

    if field_key == "garantia":
        g = specs.get("garantia") or {}
        if g.get("meses"):
            return ("✓ Listo", f"{g['meses']} meses (OEM)", g.get("fuente", ""), "alta",
                    f"retailer marketing: {g.get('meses_marketing', '?')}m — discrepancia")
        return ("✗ Falta", "", "", "", "")

    if field_key == "nom":
        n = specs.get("nom") or {}
        if n.get("certificada") in ("si",):
            return ("✓ Listo", ",".join(n.get("aplicables", [])), "", "alta", n.get("notas", ""))
        return ("✗ Falta", n.get("certificada", "desconocido"), "", "baja", n.get("notas", ""))

    if field_key == "pais_origen":
        po = specs.get("pais_origen")
        if po and po != "desconocido":
            return ("✓ Listo", po, specs.get("pais_origen_fuente", ""), "alta", "")
        return ("✗ Falta", "", "", "", "")

    # COMERCIAL
    if field_key == "categoria_shopify":
        t = com.get("taxonomia_shopify") or {}
        if t.get("ruta_completa"):
            return ("✓ Listo", t["ruta_completa"], "", "alta", t.get("razonamiento", "")[:120])
        return ("✗ Falta", "", "", "", "")

    if field_key == "tags":
        tags = com.get("tags") or []
        if len(tags) >= 5:
            return ("✓ Listo", f"{len(tags)} tags", "", "alta", ", ".join(tags[:5]) + "...")
        if tags:
            return ("⚠ Parcial", f"{len(tags)} tags", "", "media", "")
        return ("✗ Falta", "", "", "", "")

    if field_key == "precio_duarte":
        # Costo F24 viene del xlsx (col E, lo que Sergio cobra a F24). El precio retail final lo decide F24,
        # no es alcance del research. Solo registramos el costo + precio sugerido si comercial lo trajo.
        costo = sku_row.get("precio_costo_f24", 0)
        pv = com.get("precio_venta_sugerido_mxn")
        if costo:
            valor = f"costo F24 ${costo:,.2f}"
            if pv:
                valor += f" | sugerido ${pv:,.2f}"
            return ("✓ Listo", valor, "", "alta", "F24 decide retail final")
        return ("✗ Falta", "", "", "", "sin costo en xlsx")

    if field_key == "inventario":
        bodega = sku_row.get("bodega", "")
        if bodega:
            return ("✓ Listo", f"Bodega {bodega}", "", "alta", "")
        return ("✗ Falta", "", "", "", "")

    # Copy fields (los llenará el agente COPY después)
    if field_key in ("desc_corta", "desc_larga", "bullets", "seo"):
        copy_md = sku_row.get("_folder", ROOT) / "05. COPY" / "copy.md"
        if copy_md.exists() and copy_md.stat().st_size > 100:
            return ("✓ Listo", f"ver 05. COPY/copy.md ({field_key})", "", "alta", "")
        return ("✗ Falta", "", "", "", "pendiente agente COPY")

    # LOGÍSTICA
    if field_key == "peso_dim_embarque":
        e = specs.get("embarque") or {}
        if e.get("peso_kg") and e.get("dim_largo_cm"):
            return ("✓ Listo",
                    f"{e['peso_kg']}kg / {e['dim_largo_cm']}x{e['dim_ancho_cm']}x{e['dim_alto_cm']}cm",
                    "", "media", e.get("notas", "")[:120])
        return ("✗ Falta", "", "", "", "")

    if field_key == "fragil":
        e = specs.get("embarque") or {}
        if e.get("fragil") and e.get("fragil") != "desconocido":
            flete = e.get("flete_especial", "?")
            return ("✓ Listo", f"frágil={e['fragil']}, flete={flete}", "", "alta", e.get("notas", "")[:120])
        return ("✗ Falta", "", "", "", "")

    return ("✗ Falta", "", "", "", "no mapeado")


def collect_asset_index(sku: str, folder: Path, findings: dict) -> list[dict]:
    """Scan disk + JSON sources to build Asset Index rows."""
    rows = []
    media = findings.get("media") or {}
    docs = findings.get("docs") or {}

    # MEDIA
    for fld_key, tipo in [("hero", "hero"), ("secundarias", "secundaria"),
                          ("lifestyle", "lifestyle")]:
        fld = (media.get("fields") or {}).get(fld_key, {})
        for f in fld.get("files", []):
            fname = f.get("filename")
            if not fname:
                continue
            # Find file on disk
            disk_path = folder / "01. MEDIA" / (
                "hero" if fld_key == "hero" else
                "secundarias" if fld_key == "secundarias" else "lifestyle"
            ) / fname
            size_kb = round(disk_path.stat().st_size / 1024, 1) if disk_path.exists() else f.get("size_kb", 0)
            rel = disk_path.relative_to(folder).as_posix() if disk_path.exists() else f"01. MEDIA/{fld_key}/{fname}"
            rows.append({
                "sku": sku, "tipo": tipo, "filename": fname,
                "ruta": rel, "size_kb": size_kb,
                "source_url": f.get("source_url", ""),
                "notas": f.get("notas", "")
            })

    # VIDEO urls
    fld = (media.get("fields") or {}).get("video", {})
    for v in fld.get("urls", []):
        rows.append({
            "sku": sku, "tipo": "video_url", "filename": v.get("title", "video")[:60],
            "ruta": "01. MEDIA/video/video_urls.md", "size_kb": 0,
            "source_url": v.get("url", ""),
            "notas": f"{v.get('type','')} | {v.get('duration_seconds','?')}s | {v.get('language','')}"
        })

    # DOCS
    for fld_key, tipo in [("ficha_tecnica", "ficha_tecnica"),
                          ("manual", "manual"), ("exploded", "exploded")]:
        fld = (docs.get("fields") or {}).get(fld_key, {})
        if fld.get("status") != "found":
            continue
        fname = fld.get("filename", "")
        disk_path = folder / "02. DOCS" / fname
        size_kb = round(disk_path.stat().st_size / 1024, 1) if disk_path.exists() else fld.get("size_kb", 0)
        rows.append({
            "sku": sku, "tipo": tipo, "filename": fname,
            "ruta": f"02. DOCS/{fname}", "size_kb": size_kb,
            "source_url": fld.get("source_url", ""),
            "notas": f"pages={fld.get('pages','?')} lang={fld.get('language','?')}"
        })

    # Backup PDF if any extra exists
    for extra in (folder / "02. DOCS").glob("*.pdf"):
        if extra.name not in [r["filename"] for r in rows]:
            rows.append({
                "sku": sku, "tipo": "ficha_alt", "filename": extra.name,
                "ruta": f"02. DOCS/{extra.name}",
                "size_kb": round(extra.stat().st_size / 1024, 1),
                "source_url": "", "notas": "backup alterno"
            })

    return rows


def get_sku_row(wb, sku: str) -> dict:
    """
    IMPORTANTE — convención de columnas validada por Gibran 2026-05-11:
    - Col E "Precio dist Marvelsa" (row[4]) = precio que SERGIO COBRA A FERRE24 = costo F24
    - Col F "Precio neto Sergio" (row[5]) = costo interno de Sergio desde Marvelsa (referencia, NO usar para pricing)
    Inicialmente lo interpreté al revés y el COPY salió por debajo del costo (-$1,358/venta).
    """
    ws = wb["50 SKUs Prioritarios"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == sku:
            return {
                "sku": row[0],
                "descripcion": row[1],
                "url_marvelsa": row[10],
                "precio_costo_f24": row[4],       # col E — costo F24 (lo que Sergio nos cobra)
                "precio_sergio_interno": row[5],  # col F — costo interno Sergio (referencia)
                "categoria_master": row[7],
                "bodega": row[12],
            }
    return {}


def upsert_research_log(wb, sku: str, rows: list[tuple]):
    """rows = list of (campo, valor, fuente_url, confidence, agente, notas)"""
    ws = wb["Research Log"]
    # Remove existing rows for this SKU
    to_delete = [r[0].row for r in ws.iter_rows(min_row=2) if r[0].value == sku]
    for ridx in reversed(to_delete):
        ws.delete_rows(ridx)
    for (campo, valor, fuente, conf, agente, notas) in rows:
        ws.append([sku, campo, valor, fuente, conf, agente, TODAY, notas])


def upsert_asset_index(wb, sku: str, rows: list[dict]):
    ws = wb["Asset Index"]
    to_delete = [r[0].row for r in ws.iter_rows(min_row=2) if r[0].value == sku]
    for ridx in reversed(to_delete):
        ws.delete_rows(ridx)
    for r in rows:
        ws.append([r["sku"], r["tipo"], r["filename"], r["ruta"],
                   r["size_kb"], r["source_url"], TODAY, r["notas"]])


def update_checklist(wb, sku: str, states: dict):
    """states = {field_key: (flag, ...)}"""
    ws = wb["Checklist Recursos"]
    target_row = None
    for ri, row in enumerate(ws.iter_rows(min_row=3, values_only=False), start=3):
        if row[0].value == sku:
            target_row = ri
            break
    if not target_row:
        print(f"  WARN: SKU {sku} not found in Checklist")
        return

    # Update each field column
    ok_count = 0
    partial_count = 0
    for fkey, fname, col in CHECKLIST_FIELDS:
        flag = states[fkey][0]
        cell = ws.cell(row=target_row, column=col, value=flag)
        if flag.startswith("✓"):
            cell.fill = FILL_OK
            ok_count += 1
        elif flag.startswith("⚠"):
            cell.fill = FILL_PARTIAL
            partial_count += 1
        else:
            cell.fill = FILL_MISS

    # % Completo (col 4): ✓=1.0, ⚠=0.5, ✗=0.0 → sum / 24
    pct = (ok_count + 0.5 * partial_count) / len(CHECKLIST_FIELDS)
    ws.cell(row=target_row, column=4, value=pct).number_format = "0.00%"
    # Notas (col 5)
    ws.cell(row=target_row, column=5, value=f"Updated {TODAY}: {ok_count} OK, {partial_count} parcial, {len(CHECKLIST_FIELDS)-ok_count-partial_count} falta")
    print(f"  Checklist updated: {ok_count} OK, {partial_count} parcial, %completo={pct:.1%}")


def main():
    sku = sys.argv[1] if len(sys.argv) > 1 else "HP5.5N"
    folder_name = sys.argv[2] if len(sys.argv) > 2 else "Hidrolavadora a gasolina 5.5hp"
    folder = ROOT / folder_name

    if not folder.exists():
        print(f"FATAL: folder {folder} not found")
        sys.exit(1)

    findings = load_findings(folder)
    wb = openpyxl.load_workbook(XLSX)
    sku_row = get_sku_row(wb, sku)
    if not sku_row:
        print(f"FATAL: SKU {sku} not in 50 SKUs Prioritarios")
        sys.exit(1)
    sku_row["_folder"] = folder

    # Compute per-field state
    states = {}
    research_rows = []
    AGENT_BY_FIELD = {
        **{k: "MEDIA" for k in ("hero", "secundarias", "lifestyle", "video")},
        **{k: "DOCS" for k in ("ficha_tecnica", "manual", "exploded")},
        **{k: "SPECS" for k in ("marca_modelo", "motor", "operativas", "fisicas",
                                "garantia", "nom", "pais_origen", "peso_dim_embarque", "fragil")},
        **{k: "COMERCIAL" for k in ("categoria_shopify", "tags", "precio_duarte", "inventario")},
        **{k: "COPY" for k in ("desc_corta", "desc_larga", "bullets", "seo")},
    }
    for fkey, fname, col in CHECKLIST_FIELDS:
        flag, valor, fuente, conf, notas = derive_field_state(fkey, findings, sku_row)
        states[fkey] = (flag, valor, fuente, conf, notas)
        research_rows.append((fname, valor, fuente, conf, AGENT_BY_FIELD.get(fkey, ""), notas))

    asset_rows = collect_asset_index(sku, folder, findings)

    upsert_research_log(wb, sku, research_rows)
    upsert_asset_index(wb, sku, asset_rows)
    update_checklist(wb, sku, states)

    wb.save(XLSX)
    print(f"Saved {XLSX.name}")
    print(f"  Research Log: {len(research_rows)} rows for {sku}")
    print(f"  Asset Index: {len(asset_rows)} rows for {sku}")


if __name__ == "__main__":
    main()
